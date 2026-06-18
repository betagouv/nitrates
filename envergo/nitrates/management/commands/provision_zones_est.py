"""Provisionne les zones Est 1 / Est 2 (PAR7 Grand Est) depuis l'Excel juriste.

Source de verite : `envergo/nitrates/specs/zones_est_par_grand_est.xlsx`
(committe dans le repo, envoye par Karine, derive de l'arrete PAR7 consolide
Grand Est 2025 - Article 3 + Annexes 1/2).

La commande parse les onglets `Zone Est 1` et `Zone Est 2` de l'Excel et
(re)genere un CSV plat `assets/zones_est_grand_est.csv`. Ce CSV est ensuite
lu au runtime via `zonage_zones_est._mapping()` (meme pattern que
`zonage_montagne`, pas de PostGIS, pas de DB : resolution sur code INSEE).

  Zone Est 1 (alinea 1 art. 3) : allongement interdiction Type II/III sur
    mais et prairies>6mois / luzerne. 720 communes listees explicitement
    (08/51/52/57) + departements entiers 54/55/88 (toutes communes ZV).
  Zone Est 2 (alinea 2 art. 3) : allongement Type II/III pour la VIGNE
    uniquement. 4 departements entiers (08/10/51/52), pas d'annexe commune.

Recouvrement assume : 08/51/52 sont dans les deux zones. Une commune peut
etre Est 1 (mais/prairie) ET Est 2 (vigne) -> deux flags distincts.

Idempotent : la commande reecrit integralement le CSV a chaque run. Deux
runs successifs sur le meme Excel produisent un CSV byte-identique (tri
deterministe). C'est l'equivalent CSV du `update_or_create` des imports SIG.

Usage :

    docker compose run --rm django python manage.py provision_zones_est

    # Avec un Excel custom :
    docker compose run --rm django python manage.py provision_zones_est \\
        --file /path/to/zones_est.xlsx

    # Verifier sans ecrire (CI / pre-commit) :
    docker compose run --rm django python manage.py provision_zones_est --check
"""

import csv
import io
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError

APP_DIR = Path(__file__).resolve().parents[2]
DEFAULT_XLSX = APP_DIR / "specs" / "zones_est_par_grand_est.xlsx"
CSV_PATH = APP_DIR / "assets" / "zones_est_grand_est.csv"

CSV_FIELDS = ("zone", "code_departement", "code_insee", "portee")

# Mapping onglet Excel -> identifiant de zone dans le CSV genere.
SHEETS = {
    "Zone Est 1": "est_1",
    "Zone Est 2": "est_2",
}

# Valeur `portee` de l'Excel = "commune" ou commence par "TOUT LE DEPARTEMENT".
PORTEE_COMMUNE = "commune"
PORTEE_DEPARTEMENT = "departement"


class Command(BaseCommand):
    help = "Genere assets/zones_est_grand_est.csv depuis l'Excel des zones Est."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            type=Path,
            default=DEFAULT_XLSX,
            help=(
                "Chemin vers l'Excel des zones Est "
                f"(defaut : {DEFAULT_XLSX.relative_to(APP_DIR.parent.parent)})."
            ),
        )
        parser.add_argument(
            "--check",
            action="store_true",
            help=(
                "Ne reecrit pas le CSV ; sort en erreur si le CSV genere "
                "differe du CSV committe (utile en CI / pre-commit)."
            ),
        )

    def handle(self, *args, **options):
        xlsx_path: Path = options["file"]
        if not xlsx_path.exists():
            raise CommandError(f"Excel introuvable : {xlsx_path}")

        rows = self._parse(xlsx_path)
        new_csv = self._render_csv(rows)

        if options["check"]:
            current = CSV_PATH.read_text(encoding="utf-8") if CSV_PATH.exists() else ""
            if current != new_csv:
                raise CommandError(
                    f"{CSV_PATH.name} est perime vs {xlsx_path.name}. "
                    f"Relancer `manage.py provision_zones_est`."
                )
            self.stdout.write(self.style.SUCCESS(f"{CSV_PATH.name} a jour."))
            return

        CSV_PATH.parent.mkdir(parents=True, exist_ok=True)
        CSV_PATH.write_text(new_csv, encoding="utf-8")

        communes = sum(1 for r in rows if r["portee"] == PORTEE_COMMUNE)
        depts = sum(1 for r in rows if r["portee"] == PORTEE_DEPARTEMENT)
        self.stdout.write(
            self.style.SUCCESS(
                f"OK : {CSV_PATH.name} ecrit ({len(rows)} regles : "
                f"{communes} communes, {depts} departements entiers)."
            )
        )

    # ─── Parsing Excel ─────────────────────────────────────────────────────

    def _parse(self, xlsx_path: Path) -> list[dict]:
        """Lit les onglets Zone Est 1 / Zone Est 2, retourne une liste de
        regles {zone, code_departement, code_insee, portee}.

        Une regle a portee=commune porte un code_insee ; une regle
        portee=departement a un code_insee vide (le resolveur matchera sur
        le seul code_departement)."""
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        try:
            for sheet in SHEETS:
                if sheet not in wb.sheetnames:
                    raise CommandError(
                        f"Onglet manquant dans {xlsx_path.name} : {sheet!r}. "
                        f"Onglets trouves : {wb.sheetnames}"
                    )

            rules: list[dict] = []
            for sheet, zone in SHEETS.items():
                rules.extend(self._parse_sheet(wb[sheet], zone, xlsx_path))
        finally:
            wb.close()

        # Dedup (au cas ou l'Excel contient des doublons) + tri deterministe
        # pour que deux runs produisent un CSV byte-identique.
        seen = set()
        deduped = []
        for r in rules:
            key = (r["zone"], r["code_departement"], r["code_insee"])
            if key in seen:
                continue
            seen.add(key)
            deduped.append(r)
        deduped.sort(key=lambda r: (r["zone"], r["code_departement"], r["code_insee"]))
        return deduped

    def _parse_sheet(self, ws, zone: str, xlsx_path: Path) -> list[dict]:
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            return []
        cols = {str(name).strip(): i for i, name in enumerate(header) if name}
        for required in ("code_departement", "portee"):
            if required not in cols:
                raise CommandError(
                    f"Colonne {required!r} absente de l'onglet {ws.title!r} "
                    f"({xlsx_path.name}). Colonnes : {list(cols)}"
                )

        out = []
        for raw in rows:
            if raw is None or all(c is None for c in raw):
                continue
            code_dep = _cell(raw, cols, "code_departement")
            portee_raw = _cell(raw, cols, "portee")
            code_insee = _cell(raw, cols, "code_insee")
            if not code_dep:
                continue

            if portee_raw.lower().startswith("commune"):
                if not code_insee:
                    self.stderr.write(
                        f"  [warn] {ws.title}: ligne commune sans code_insee "
                        f"(dep {code_dep}), ignoree."
                    )
                    continue
                out.append(
                    {
                        "zone": zone,
                        "code_departement": code_dep,
                        "code_insee": code_insee,
                        "portee": PORTEE_COMMUNE,
                    }
                )
            else:
                # "TOUT LE DEPARTEMENT (...)" -> regle au departement entier.
                out.append(
                    {
                        "zone": zone,
                        "code_departement": code_dep,
                        "code_insee": "",
                        "portee": PORTEE_DEPARTEMENT,
                    }
                )
        return out

    # ─── Rendu CSV ─────────────────────────────────────────────────────────

    def _render_csv(self, rows: list[dict]) -> str:
        buf = io.StringIO(newline="")
        writer = csv.DictWriter(buf, fieldnames=CSV_FIELDS, lineterminator="\n")
        writer.writeheader()
        for r in rows:
            writer.writerow(r)
        return buf.getvalue()


def _cell(raw, cols, name) -> str:
    idx = cols.get(name)
    if idx is None or idx >= len(raw):
        return ""
    val = raw[idx]
    return "" if val is None else str(val).strip()
